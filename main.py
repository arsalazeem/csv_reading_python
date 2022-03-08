import openpyxl
import json
import csv
from csv import reader as rd

# book = openpyxl.load_workbook('file_1.xlsx')
# sheet = book.active
# k=0
# for i in range(1,1000):
#     a1 = sheet.cell(row=i, column=1)
#     if type(a1.value)==int:
#         sellerId=sheet.cell(row=i,column=1)
#         sellerId=sellerId.value
#         email = sheet.cell(row=i, column=2)
#         email= email.value
#         serviceId= sheet.cell(row=i, column=3)
#         serviceId = serviceId.value
#         serviceName= sheet.cell(row=i, column=3)
#         serviceName = serviceName.value
#         description = sheet.cell(row=i, column=3)
#         description = description.value
#         isFeatured = sheet.cell(row=i, column=3)
#         isFeatured = isFeatured.value
#         categoryName = sheet.cell(row=i, column=3)
#         categoryName= categoryName.value

count = 0
master_list = []
# file = open("activeSellers.csv",encoding='cp1252')
# reader = csv.reader(file)
# lines= len(list(reader))
# print(lines)
# quit()
with open('second_file.csv', 'r', encoding='cp1252') as file:
    reader = csv.reader(file)
    try:

        for row in reader:
            print(row)
            count = count + 1
            try:
                current_id = int(row[0])
                selleId = row[0]
                firstName = row[1]
                lastName = row[2]
                email = row[3]
                dialCode = row[4]
                phoneNumber = row[5]
                aboutMe = row[6]
                sellerType = row[7]
                agencyName = row[8]
                title = row[9]
                streetAddress = row[10]
                apartment = row[11]
                city = row[12]
                state = row[13]
                master_list.append({
                    "selleId": current_id,
                    "firstName": firstName,
                    "lastName": lastName,
                    "email": email,
                    "dialCode": dialCode,
                    "phoneNumber": phoneNumber,
                    "aboutMe": aboutMe,
                    "sellerType": sellerType,
                    "agencyName": agencyName,
                    "title": title,
                    "streetAddress": streetAddress,
                    "apartment": apartment,
                    "city": city,
                    "state": state})

            except:
                pass
    except:
        pass


jsonString = json.dumps(master_list)
jsonFile = open("activeSellers.json", "w")
jsonFile.write(jsonString)
jsonFile.close()
# import csv
# rdr = csv.reader(open("second_file.csv"))
# line1 = rd().__next__()
# line2 = rdr.__next__()
# second_file
# firstName
# lastName
# email
# dialCode
# phoneNumber
# aboutMe
# sellerType
# agencyName
# title
# streetAddress
# apartment
# city
# state
# zipCode
# endofsecondfile
# sellerId
# email
# serviceId
# serviceName
# description
# isFeatured
# categoryId
# categoryName
# packageId
# packageType
# pDescription
# deliveryTimeInDays
# isSourceFileIncluded
# price
# revisions
# second_file
# firstName
# lastName
# email
# dialCode
# phoneNumber
# aboutMe
# sellerType
# agencyName
# title
# streetAddress
# apartment
# city
# state
# zipCode
# print(count)
# jsonString = json.dumps(master_list)
# jsonFile = open("activeSellers.json", "w")
# jsonFile.write(jsonString)
# jsonFile.close()
# f = open("file1.txt", "r")
# i=0
# while True:
#     try:
#         print(f.read(i))
#         i = i + 1
#     except:
#         break
